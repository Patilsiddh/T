import os
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
from flask_mail import Mail, Message




# ✅ Detect environment (Vercel vs Local)
if os.getenv("VERCEL"):
    BASE_DIR = "/tmp"
else:
    BASE_DIR = "."

# ---------------- Paths ---------------- #
DB_FOLDER = os.path.join(BASE_DIR, "database")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static/uploads")

DB_PATH = os.path.join(DB_FOLDER, "Tataplay.db")
EXCEL_PATH = os.path.join(DB_FOLDER, "user_logs.xlsx")

# ✅ Create folders (works in /tmp)
os.makedirs(DB_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------- Excel Logging ---------------- #

EXCEL_FILE = "contacts.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        ws.append(["Name", "Mobile", "Email", "Message", "Date & Time"])

        wb.save(EXCEL_FILE)


def save_to_excel(name, mobile, email, message):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["Name", "Mobile", "Email", "Message", "Date & Time"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        name,
        mobile,
        email,
        message,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(EXCEL_FILE) 
         
def log_user_action(username, action):
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UserLogs"
        ws.append([ "Login"])
        wb.save(EXCEL_PATH)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    row = ws.max_row + 1
    
    ws.cell(row=row, column=2, value=username if action == "login" else "")
    wb.save(EXCEL_PATH)

# ---------------- Flask App ---------------- #
app = Flask(__name__)
app.secret_key = "supersecretkey"  # Use environment variable in production
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# EMAIL CONFIG (Gmail example)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'patilsiddh2026@gmail.com'
app.config['MAIL_PASSWORD'] = 'egpzzvflxvkpibyj'  # NOT normal password

mail = Mail(app)
import os
import sqlite3
from werkzeug.security import generate_password_hash

# ---------------- Database ---------------- #
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():

    # =========================
    # DROP OLD DB AUTOMATICALLY
    # =========================
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print("Old database deleted and recreated")

    conn = get_db_connection()
    c = conn.cursor()

    # USERS
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            is_admin INTEGER DEFAULT 0
        )
    """)

    # CATEGORIES
    c.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE
        )
    """)

    # GLOBAL OFFER
    c.execute("""
        CREATE TABLE IF NOT EXISTS global_offer (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            text TEXT NOT NULL,
            active INTEGER DEFAULT 1
        )
    """)

    # PLANS
    c.execute("""
        CREATE TABLE IF NOT EXISTS plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            description TEXT,
            image_url TEXT,
            is_best_seller INTEGER DEFAULT 0,
            category_id INTEGER,
            FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE SET NULL
        )
    """)

    # OFFERS
    c.execute("""
        CREATE TABLE IF NOT EXISTS offers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT,
            subtitle TEXT,
            amount_text TEXT,
            image_url TEXT,
            offer_type TEXT,
            is_popup INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            start_date TEXT,
            end_date TEXT
        )
    """)

    # PLAN IMAGES
    c.execute("""
        CREATE TABLE IF NOT EXISTS plan_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER,
            filename TEXT,
            FOREIGN KEY(plan_id) REFERENCES plans(id) ON DELETE CASCADE
        )
    """)

    # PLAN DURATIONS
    c.execute("""
        CREATE TABLE IF NOT EXISTS plan_durations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER,
            duration INTEGER,
            FOREIGN KEY(plan_id) REFERENCES plans(id) ON DELETE CASCADE
        )
    """)

    # PLAN SPEEDS
    c.execute("""
        CREATE TABLE IF NOT EXISTS plan_speeds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            duration_id INTEGER,
            speed INTEGER,
            price REAL,
            discounted_price REAL DEFAULT 0,
            FOREIGN KEY(duration_id) REFERENCES plan_durations(id) ON DELETE CASCADE
        )
    """)

    # SERVICES
    c.execute("""
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            category TEXT,
            price REAL
        )
    """)

    # =========================
    # DEFAULT DATA
    # =========================

    c.execute("INSERT OR IGNORE INTO categories (name) VALUES ('Broadband')")
    c.execute("INSERT OR IGNORE INTO categories (name) VALUES ('DTH')")

    c.execute("SELECT id FROM categories WHERE name='Broadband'")
    broadband_id = c.fetchone()[0]

    c.execute("SELECT id FROM categories WHERE name='DTH'")
    dth_id = c.fetchone()[0]

    durations = [1, 2, 3, 6, 12]

    # =========================
    # BROADBAND PLANS
    # =========================
    broadband_plans = [
        ("Basic Fiber", "Perfect for browsing & YouTube HD streaming"),
        ("Smart Fiber", "Ideal for families & OTT streaming"),
        ("Ultra Fiber", "4K streaming + gaming + heavy usage")
    ]

    duration_prices = {
        1: 100,
        2: 500,
        3: 300,
        6: 600,
        12: 1000
    }

    for name, desc in broadband_plans:
        c.execute("""
            INSERT INTO plans (name, description, category_id, is_best_seller)
            VALUES (?, ?, ?, ?)
        """, (name, desc, broadband_id, 1))

        plan_id = c.lastrowid

        for d in durations:
            price = duration_prices[d]

            c.execute("""
                INSERT INTO plan_durations (plan_id, duration)
                VALUES (?, ?)
            """, (plan_id, d))

            dur_id = c.lastrowid

            c.execute("""
                INSERT INTO plan_speeds (duration_id, speed, price, discounted_price)
                VALUES (?, ?, ?, ?)
            """, (dur_id, 0, price, price))

    # =========================
    # DTH PLANS
    # =========================
    dth_plans = [
        ("Basic DTH Pack", "Regional channels + HD support", 199),
        ("Family DTH Pack", "Entertainment + Movies + Sports", 299),
        ("Premium DTH Pack", "All HD channels + OTT included", 499)
    ]

    for name, desc, price in dth_plans:
        c.execute("""
            INSERT INTO plans (name, description, category_id, is_best_seller)
            VALUES (?, ?, ?, ?)
        """, (name, desc, dth_id, 1))

        plan_id = c.lastrowid

        for d in durations:
            c.execute("""
                INSERT INTO plan_durations (plan_id, duration)
                VALUES (?, ?)
            """, (plan_id, d))

            dur_id = c.lastrowid

            c.execute("""
                INSERT INTO plan_speeds (duration_id, speed, price, discounted_price)
                VALUES (?, ?, ?, ?)
            """, (dur_id, 0, price * d, price * d))

    # =========================
    # OFFERS
    # =========================
    c.execute("""
        INSERT OR IGNORE INTO offers (title, subtitle, amount_text, offer_type, is_popup, is_active)
        VALUES ('🔥 Broadband Offer', 'Unlimited High Speed Fiber', '₹799/month', 'broadband', 0, 1)
    """)

    c.execute("""
        INSERT OR IGNORE INTO offers (title, subtitle, amount_text, offer_type, is_popup, is_active)
        VALUES ('📺 DTH Special Offer', 'HD Channels + Free Installation', '₹499/month', 'dth', 0, 1)
    """)

    c.execute("""
        INSERT OR IGNORE INTO offers (title, subtitle, amount_text, offer_type, is_popup, is_active)
        VALUES ('⚡ Limited Time Offer', 'Get 50% OFF on first 3 months', '₹299 only', 'popup', 1, 1)
    """)

    # GLOBAL OFFER
    c.execute("""
        INSERT OR IGNORE INTO global_offer (text, active)
        VALUES ('🔥 Tata Play Fiber: Best Broadband + DTH Combo Offers Available Now!', 1)
    """)

    # ADMIN
    admin_username = "admin"
    admin_password = generate_password_hash("admin123")

    c.execute("SELECT * FROM users WHERE username=?", (admin_username,))
    if not c.fetchone():
        c.execute(
            "INSERT INTO users (username, password, is_admin) VALUES (?,?,?)",
            (admin_username, admin_password, 1)
        )

    conn.commit()
    conn.close()


init_db()
# ---------------- Routes ---------------- #
@app.route('/')
def home():
    conn = get_db_connection()

    # ✅ FETCH ACTIVE OFFERS
    broadband_offer = conn.execute("""
        SELECT * FROM offers 
        WHERE offer_type='broadband' AND is_active=1
        LIMIT 1
    """).fetchone()

    dth_offer = conn.execute("""
        SELECT * FROM offers 
        WHERE offer_type='dth' AND is_active=1
        LIMIT 1
    """).fetchone()

    # ✅ THIS IS YOUR POPUP QUERY (PUT HERE)
    popup_offer = conn.execute("""
        SELECT * FROM offers 
        WHERE offer_type='popup' AND is_active=1
        LIMIT 1
    """).fetchone()

    conn.close()

    # ✅ PASS TO TEMPLATE
    return render_template(
        'index.html',
        broadband_offer=broadband_offer,
        dth_offer=dth_offer,
        popup_offer=popup_offer
    )

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT * FROM users WHERE username=?', (username,))
        user = c.fetchone()
        conn.close()

        # ✅ Allow ONLY admin login
        if user and check_password_hash(user['password'], password):
            if user['is_admin'] != 1:
                flash('Access denied! Admins only.', "error")
                return redirect(url_for('login'))

            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = user['is_admin']

            log_user_action(username, "login")
            return redirect(url_for('admin'))

        else:
            flash('Invalid credentials!', "error")

    return render_template('login.html')
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))





# ---------------- Plan Routes ---------------- #
@app.route('/edit_plan/<int:id>', methods=['GET', 'POST'])
def edit_plan(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # ===== POST method: update plan =====
    if request.method == 'POST':
        # Update main plan details
        name = request.form['plan_name']
        description = request.form.get('description', '')
        is_best_seller = 1 if request.form.get('best_seller') else 0

        cursor.execute("""
            UPDATE plans 
            SET name=?, description=?, is_best_seller=? 
            WHERE id=?
        """, (name, description, is_best_seller, id))

        # Fetch existing durations for this plan
        cursor.execute("SELECT id FROM plan_durations WHERE plan_id=?", (id,))
        durations = cursor.fetchall()

        for dur in durations:
            dur_id = dur['id']

            # Collect dynamic fields for this duration
            speeds = request.form.getlist(f"speed_{dur_id}[]")
            prices = request.form.getlist(f"price_{dur_id}[]")
            discounted_prices = request.form.getlist(f"discounted_price_{dur_id}[]")

            # Fetch existing speeds from DB
            cursor.execute("SELECT id FROM plan_speeds WHERE duration_id=?", (dur_id,))
            existing_speeds = cursor.fetchall()

            # Update existing speeds or insert new ones
            for i, (spd, prc, dprc) in enumerate(zip(speeds, prices, discounted_prices)):
                spd = int(spd) if spd else 0
                prc = float(prc) if prc else 0.0
                dprc = float(dprc) if dprc else 0.0

                if i < len(existing_speeds):
                    cursor.execute("""
                        UPDATE plan_speeds 
                        SET speed=?, price=?, discounted_price=? 
                        WHERE id=?
                    """, (spd, prc, dprc, existing_speeds[i]['id']))
                else:
                    cursor.execute("""
                        INSERT INTO plan_speeds (duration_id, speed, price, discounted_price)
                        VALUES (?, ?, ?, ?)
                    """, (dur_id, spd, prc, dprc))

            # Delete extra speeds if removed in form
            if len(existing_speeds) > len(speeds):
                ids_to_delete = [s['id'] for s in existing_speeds[len(speeds):]]
                cursor.execute(
                    f"DELETE FROM plan_speeds WHERE id IN ({','.join('?'*len(ids_to_delete))})",
                    ids_to_delete
                )

        # ===== Commit & close =====
        conn.commit()
        conn.close()
        flash("Plan updated successfully!", "success")
        return redirect(url_for('admin'))

    # ===== GET method: show plan for editing =====
    plan = cursor.execute("SELECT * FROM plans WHERE id=?", (id,)).fetchone()

    # Fetch durations and speeds for form
    durations = cursor.execute("SELECT * FROM plan_durations WHERE plan_id=?", (id,)).fetchall()
    plan_durations = []
    for dur in durations:
        dur_id = dur['id']
        speeds = cursor.execute("SELECT * FROM plan_speeds WHERE duration_id=?", (dur_id,)).fetchall()
        plan_durations.append({
            'duration': dur['duration'],
            'id': dur_id,
            'speeds': speeds
        })

    conn.close()
    return render_template('edit_plan.html', plan=plan, plan_durations=plan_durations)

# ---------------- Service Routes ---------------- #
@app.route('/add_service', methods=['POST'])
def add_service():
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    name = request.form['name']
    category = request.form['category']
    price = float(request.form['price'])
    conn = get_db_connection()
    conn.execute("INSERT INTO services (name, category, price) VALUES (?,?,?)", (name, category, price))
    conn.commit()
    conn.close()
    flash("Service added!", "success")
    return redirect(url_for('admin'))

@app.route('/edit_service/<int:id>', methods=['POST'])
def edit_service(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    name = request.form['name']
    category = request.form['category']
    price = float(request.form['price'])
    conn = get_db_connection()
    conn.execute("UPDATE services SET name=?, category=?, price=? WHERE id=?", (name, category, price, id))
    conn.commit()
    conn.close()
    flash("Service updated!", "success")
    return redirect(url_for('admin'))


@app.route('/delete_service/<int:id>')
def delete_service(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    conn.execute("DELETE FROM services WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash("Service deleted!", "success")
    return redirect(url_for('admin'))

# ---------------- Category Routes ---------------- #
@app.route('/add_category', methods=['POST'])
def add_category():
    name = request.form.get('new_category')
    if not name:
        flash('Category name cannot be empty!', "error")
        return redirect(url_for('admin'))
    try:
        conn = get_db_connection()
        conn.execute("INSERT INTO categories (name) VALUES (?)", (name,))
        conn.commit()
        conn.close()
        flash(f'Category "{name}" added!', "success")
    except sqlite3.IntegrityError:
        flash(f'Category "{name}" already exists!', "error")
    return redirect(url_for('admin'))

@app.route('/edit_category/<int:id>', methods=['POST'])
def edit_category(id):
    new_name = request.form.get('category_name')
    if not new_name:
        flash('Category name cannot be empty!', "error")
        return redirect(url_for('admin'))
    try:
        conn = get_db_connection()
        conn.execute("UPDATE categories SET name=? WHERE id=?", (new_name, id))
        conn.commit()
        conn.close()
        flash(f'Category updated to "{new_name}"!', "success")
    except sqlite3.IntegrityError:
        flash(f'Category "{new_name}" already exists!', "error")
    return redirect(url_for('admin'))


@app.route('/update_offer_selection', methods=['POST'])
def update_offer_selection():

    if not session.get('is_admin'):
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor()

    sections = ['broadband', 'dth', 'popup']

    for sec in sections:

        selected_id = request.form.get(f'active_{sec}')

        if selected_id:

            # ❗ First reset all offers of this type
            cursor.execute("""
                UPDATE offers
                SET is_active = 0
                WHERE offer_type = ?
            """, (sec,))

            # ❗ Then activate selected one
            cursor.execute("""
                UPDATE offers
                SET is_active = 1
                WHERE id = ?
            """, (selected_id,))

    conn.commit()
    conn.close()

    flash("Offer selection updated successfully!", "success")
    return redirect(url_for('admin'))

@app.route('/update-full-plan/<int:plan_id>', methods=['POST'])
def update_full_plan(plan_id):

    if not session.get('is_admin'):
        return jsonify({'success': False, 'msg': 'Unauthorized'})

    import os
    import json
    from werkzeug.utils import secure_filename

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # ---------------- BASIC DATA ----------------
        name = request.form.get('name')
        description = request.form.get('description')
        category_id = request.form.get('category_id')
        is_best = 1 if request.form.get('is_best_seller') == "1" else 0

        # ---------------- UPDATE PLAN ----------------
        if category_id:
            cursor.execute("""
                UPDATE plans
                SET name=?, description=?, category_id=?, is_best_seller=?
                WHERE id=?
            """, (name, description, category_id, is_best, plan_id))
        else:
            cursor.execute("""
                UPDATE plans
                SET name=?, description=?, is_best_seller=?
                WHERE id=?
            """, (name, description, is_best, plan_id))

        # ---------------- DELETE OLD DATA ----------------
        cursor.execute("""
            DELETE FROM plan_speeds
            WHERE duration_id IN (
                SELECT id FROM plan_durations WHERE plan_id=?
            )
        """, (plan_id,))

        cursor.execute("DELETE FROM plan_durations WHERE plan_id=?", (plan_id,))

        # ---------------- INSERT NEW DURATIONS + SPEEDS ----------------
        durations = json.loads(request.form.get('durations', '[]'))

        for d in durations:
            cursor.execute("""
                INSERT INTO plan_durations (plan_id, duration)
                VALUES (?, ?)
            """, (plan_id, d.get('duration')))

            duration_id = cursor.lastrowid

            # IMPORTANT FIX: loop speeds correctly
            for s in d.get('speeds', []):
                cursor.execute("""
                    INSERT INTO plan_speeds (duration_id, speed, price, discounted_price)
                    VALUES (?, ?, ?, ?)
                """, (
                    duration_id,
                    s.get('speed'),
                    s.get('price'),
                    s.get('discounted_price') or s.get('discount') or 0
                ))

        # ---------------- DELETE OLD IMAGES (OPTIONAL BUT RECOMMENDED) ----------------
        cursor.execute("DELETE FROM plan_images WHERE plan_id=?", (plan_id,))

        # ---------------- INSERT NEW IMAGES ----------------
        files = request.files.getlist('images[]')

        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)

                cursor.execute("""
                    INSERT INTO plan_images (plan_id, filename)
                    VALUES (?, ?)
                """, (plan_id, filename))

        # ---------------- COMMIT ----------------
        conn.commit()

        return jsonify({'success': True})

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)})

    finally:
        conn.close()
@app.route('/delete-plan-image/<int:image_id>', methods=['POST'])
def delete_plan_image(image_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM plan_images WHERE id=?", (image_id,))

    conn.commit()
    conn.close()

    return jsonify({"success": True})
@app.route('/add_offer', methods=['POST'])
def add_offer():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    title = request.form.get('title')
    subtitle = request.form.get('subtitle')
    amount_text = request.form.get('amount_text')
    offer_type = request.form.get('offer_type')
    is_popup = 1 if request.form.get('is_popup') else 0

    image = request.files.get('image')
    filename = None

    if image and image.filename:
        filename = secure_filename(image.filename)
        image.save(os.path.join(UPLOAD_FOLDER, filename))

    conn = get_db_connection()
    conn.execute("""
        INSERT INTO offers
        (title, subtitle, amount_text, image_url, offer_type, is_popup)
        VALUES (?,?,?,?,?,?)
    """, (title, subtitle, amount_text, filename, offer_type, is_popup))

    conn.commit()
    conn.close()

    flash("Offer added successfully!", "success")
    return redirect(url_for('admin'))

@app.route('/delete_offer/<int:id>')
def delete_offer(id):
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    conn = get_db_connection()
    conn.execute("DELETE FROM offers WHERE id=?", (id,))
    conn.commit()
    conn.close()

    flash("Offer deleted!", "success")
    return redirect(url_for('admin'))

@app.route("/support")
def support():
    return render_template("support.html")

# ---------------- Services Page ---------------- #





import threading


def send_email_async(app, msg):
    with app.app_context():
        try:
            mail.send(msg)
        except Exception as e:
            print("EMAIL ERROR:", e)


@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if request.method == 'POST':

        name = request.form.get('name')
        mobile = request.form.get('mobile')
        email = request.form.get('email')
        message = request.form.get('message')

        try:
            # ✅ FAST SAVE (no delay)
            save_to_excel(name, mobile, email, message)

            # ✅ EMAIL (NON-BLOCKING)
            msg = Message(
                subject=f"New Contact - {name}",
                sender=app.config['MAIL_USERNAME'],
                recipients=['patil123sidd@gmail.com'],  # change this
                reply_to=email
            )

            msg.body = f"""
📩 New Enquiry - Tata Play Fiber (Broadband / DTH):

Name: {name}
Mobile: {mobile}
Email: {email}
Message: {message}
"""

            # 🔥 run email in background (fix Railway timeout)
            threading.Thread(
                target=send_email_async,
                args=(app, msg)
            ).start()

            flash("✅ Message sent successfully!")
            return redirect(url_for('home'))

        except Exception as e:
            print("CONTACT ERROR:", e)
            flash("❌ Something went wrong. Please try again.")
            return redirect(url_for('contact'))

    return render_template('contact.html')



@app.route("/admin/set-offer", methods=["POST"])
def set_offer():
    if not session.get('is_admin'):
        return redirect(url_for('login'))

    text = request.form.get("offer_text", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    # ❗ ALWAYS deactivate old offers first
    cur.execute("UPDATE global_offer SET active=0")

    # ✅ ONLY insert if text is NOT empty
    if text:
        from datetime import datetime
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cur.execute("""
            INSERT INTO global_offer (text, active, updated_at)
            VALUES (?, 1, ?)
        """, (text, now))

    conn.commit()
    conn.close()

    flash("Offer updated!", "success")
    return redirect(url_for('admin'))



@app.context_processor
def inject_offer():
    conn = get_db_connection()
    cur = conn.cursor()

    offer = cur.execute("""
        SELECT text FROM global_offer
        WHERE active=1 AND text IS NOT NULL AND text != ''
        ORDER BY id DESC LIMIT 1
    """).fetchone()

    conn.close()

    return {
        "global_offer": offer["text"] if offer else None
    }


@app.route('/services')
def services():
    conn = get_db_connection()
    cur = conn.cursor()

    categories = cur.execute("SELECT * FROM categories").fetchall()

    category_data = []

    for cat in categories:
        plans = cur.execute(
            "SELECT * FROM plans WHERE category_id = ?",
            (cat["id"],)
        ).fetchall()

        plans_data = []

        for plan in plans:
            # Get durations
            durations = cur.execute(
                "SELECT * FROM plan_durations WHERE plan_id = ?",
                (plan["id"],)
            ).fetchall()

            durations_data = []

            for dur in durations:
                # Get speeds
                speeds = cur.execute(
                    "SELECT * FROM plan_speeds WHERE duration_id = ?",
                    (dur["id"],)
                ).fetchall()

                speeds_data = []
                for spd in speeds:
                    speeds_data.append({
                        "speed": spd["speed"],
                        "price": spd["price"],
                        "discounted_price": spd["discounted_price"]
                    })

                durations_data.append({
                    "duration": dur["duration"],
                    "speeds": speeds_data
                })

            # Get images
            images = cur.execute(
                "SELECT * FROM plan_images WHERE plan_id = ?",
                (plan["id"],)
            ).fetchall()

            plans_data.append({
                "plan": plan,
                "durations": durations_data,
                "images": images
            })

        category_data.append({
            "id": cat["id"],
            "name": cat["name"],
            "plans": plans_data
        })

    conn.close()

    return render_template("services.html", categories=category_data)




































@app.route('/delete_category/<int:id>')
def delete_category(id):
    conn = get_db_connection()
    conn.execute("DELETE FROM categories WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash('Category deleted!', "success")
    return redirect(url_for('admin'))
# ---------------- Admin Panel ---------------- #
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if not session.get('is_admin'):
        flash("You must be admin to access this page!", "error")
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor()
    offers = cursor.execute("SELECT * FROM offers").fetchall()

    # ---------------- ADD CATEGORY ---------------- #
    if request.method == 'POST' and 'new_category' in request.form:
        name = request.form['new_category'].strip()
        if name:
            try:
                cursor.execute("INSERT INTO categories (name) VALUES (?)", (name,))
                conn.commit()
                flash(f'Category "{name}" added!', "success")
            except sqlite3.IntegrityError:
                flash(f'Category "{name}" already exists!', "error")
        return redirect(url_for('admin'))

    # ---------------- ADD PLAN ---------------- #
    if request.method == 'POST' and 'plan_name' in request.form:
        plan_name = request.form['plan_name'].strip()
        description = request.form.get('description', '')
        category_id = request.form['category_id']
        is_best_seller = 1 if request.form.get('best_seller') else 0

        cursor.execute(
            "INSERT INTO plans (name, description, is_best_seller, category_id) VALUES (?,?,?,?)",
            (plan_name, description, is_best_seller, category_id)
        )
        plan_id = cursor.lastrowid

        # Upload Images
        files = request.files.getlist('plan_images[]')
        for file in files:
            if file.filename:
                filename = secure_filename(file.filename)
                file.save(os.path.join(UPLOAD_FOLDER, filename))
                cursor.execute("INSERT INTO plan_images (plan_id, filename) VALUES (?,?)", (plan_id, filename))

        # Durations & Speeds
        durations = request.form.getlist('duration[]')
        speeds = request.form.getlist('speed[]')
        prices = request.form.getlist('price[]')
        discounted_prices = request.form.getlist('discounted_price[]')

        speed_idx = 0
        for dur in durations:
            cursor.execute("INSERT INTO plan_durations (plan_id, duration) VALUES (?,?)", (plan_id, dur))
            dur_id = cursor.lastrowid

            s = speeds[speed_idx]
            p = prices[speed_idx]
            d = discounted_prices[speed_idx] or 0

            cursor.execute(
                "INSERT INTO plan_speeds (duration_id, speed, price, discounted_price) VALUES (?,?,?,?)",
                (dur_id, s, p, d)
            )
            speed_idx += 1

        conn.commit()
        flash(f'Plan "{plan_name}" added successfully!', "success")
        categories = cursor.execute("SELECT * FROM categories").fetchall()
        global_offer = cursor.execute("SELECT * FROM global_offer").fetchall()

        conn.close()
        return redirect(url_for('admin'))
       

    # ---------------- SHOW CATEGORIES & PLANS ---------------- #
    cursor.execute("SELECT * FROM categories")
    categories = cursor.fetchall()

    cat_list = []

    for cat in categories:
        cat_id = cat['id']

        cursor.execute("""
            SELECT p.*, 
                   GROUP_CONCAT(pi.filename) AS images
            FROM plans p
            LEFT JOIN plan_images pi ON p.id = pi.plan_id
            WHERE p.category_id=?
            GROUP BY p.id
        """, (cat_id,))
        plans = cursor.fetchall()

        plan_list = []

        for plan in plans:
            plan_id = plan['id']

            # Durations
            cursor.execute("SELECT * FROM plan_durations WHERE plan_id=?", (plan_id,))
            durations = cursor.fetchall()

            dur_list = []   # ✅ INIT HERE

            for dur in durations:
                dur_id = dur['id']

                cursor.execute("SELECT * FROM plan_speeds WHERE duration_id=?", (dur_id,))
                spds = cursor.fetchall()

                speeds_list = []
                for spd in spds:
                    speeds_list.append({
                        "id": spd["id"],
                        "speed": spd["speed"],
                        "price": spd["price"],
                        "discounted_price": spd["discounted_price"]
                    })

                # ✅ FIXED (inside loop)
                dur_list.append({
                    "id": dur["id"],
                    "duration": dur["duration"],
                    "speeds": speeds_list
                })

            # Images
            imgs = []
            if plan['images']:
                imgs = [{'filename': f} for f in plan['images'].split(',')]

            plan_list.append({
                'plan': plan,
                'images': imgs,
                'durations': dur_list
            })

        # ✅ KEEP THIS OUTSIDE plan loop
        cat_list.append({
            'id': cat_id,
            'name': cat['name'],
            'plans': plan_list
        })

    conn.close()
    return render_template(
    'admin.html',
    categories=cat_list,
    offers=offers,
    
)
@app.route('/plans')
def plans():
    conn = get_db_connection()
    cur = conn.cursor()

    categories = cur.execute("SELECT * FROM categories").fetchall()

    category_data = []

    for cat in categories:
        plans = cur.execute(
            "SELECT * FROM plans WHERE category_id = ?",
            (cat["id"],)
        ).fetchall()

        plans_data = []

        for plan in plans:

            # IMAGES
            images = cur.execute(
                "SELECT * FROM plan_images WHERE plan_id = ?",
                (plan["id"],)
            ).fetchall()

            # DURATIONS (IMPORTANT: table name = plan_durations)
            durations = cur.execute(
                "SELECT * FROM plan_durations WHERE plan_id = ?",
                (plan["id"],)
            ).fetchall()

            duration_data = []

            for dur in durations:

                # SPEEDS (IMPORTANT: table name = plan_speeds)
                speeds = cur.execute(
                    "SELECT * FROM plan_speeds WHERE duration_id = ?",
                    (dur["id"],)
                ).fetchall()

                duration_data.append({
                    "id": dur["id"],
                    "duration": dur["duration"],
                    "speeds": speeds
                })

            plans_data.append({
                "plan": plan,
                "images": images,
                "durations": duration_data
            })

        category_data.append({
            "id": cat["id"],
            "name": cat["name"],
            "plans": plans_data
        })

    conn.close()

    return render_template("plans.html", categories=category_data)
@app.route('/admin/update_plan/<int:plan_id>', methods=['POST'])
def update_plan(plan_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        import json

        data = request.form.get('json_data')
        if data:
            data = json.loads(data)
        else:
            data = {}

        # -------- UPDATE PLAN -------- #
        best = data.get('is_best_seller')

        is_best = 1 if str(best).lower() in ['true', '1', 'yes', 'on'] else 0

        cursor.execute("""
            UPDATE plans SET name=?, description=?, is_best_seller=? WHERE id=?
        """, (data.get('name'), data.get('description'), is_best, plan_id))
      

        # -------- UPDATE EXISTING DURATIONS -------- #
        for dur_id, val in data.get('duration', {}).items():
            cursor.execute(
                "UPDATE plan_durations SET duration=? WHERE id=?",
                (val, dur_id)
            )

        # -------- UPDATE EXISTING SPEEDS -------- #
        for spd_id, val in data.get('speed', {}).items():
            cursor.execute(
                "UPDATE plan_speeds SET speed=? WHERE id=?",
                (val, spd_id)
            )

        for spd_id, val in data.get('price', {}).items():
            cursor.execute(
                "UPDATE plan_speeds SET price=? WHERE id=?",
                (val, spd_id)
            )

        for spd_id, val in data.get('discounted_price', {}).items():
            cursor.execute(
                "UPDATE plan_speeds SET discounted_price=? WHERE id=?",
                (val, spd_id)
            )

        # -------- ADD NEW DURATIONS -------- #
        new_duration_map = {}

        for dur in data.get('new_durations', []):
            cursor.execute(
                "INSERT INTO plan_durations (plan_id, duration) VALUES (?, ?)",
                (plan_id, dur)
            )
            new_duration_map[dur] = cursor.lastrowid

        # -------- ADD NEW SPEEDS -------- #
        for sp in data.get('new_speeds', []):
            dur_val = sp.get('duration')

            # find duration id
            dur_id = new_duration_map.get(dur_val)

            if not dur_id:
                cursor.execute(
                    "SELECT id FROM plan_durations WHERE plan_id=? AND duration=?",
                    (plan_id, dur_val)
                )
                row = cursor.fetchone()
                if row:
                    dur_id = row["id"]

            if dur_id:
                cursor.execute("""
                    INSERT INTO plan_speeds (duration_id, speed, price, discounted_price)
                    VALUES (?, ?, ?, ?)
                """, (
                    dur_id,
                    sp.get('speed'),
                    sp.get('price'),
                    sp.get('discounted_price', 0)
                ))

        # -------- IMAGE UPLOAD -------- #
        files = request.files.getlist('images')

        for file in files:
            if file.filename:
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

                cursor.execute(
                    "INSERT INTO plan_images (plan_id, filename) VALUES (?, ?)",
                    (plan_id, filename)
                )

        # -------- DELETE IMAGES -------- #
        for img_id in data.get('delete_images', []):
            cursor.execute(
                "DELETE FROM plan_images WHERE id=?",
                (img_id,)
            )

        conn.commit()
        return jsonify({"success": True})

    except Exception as e:
        print("ERROR:", e)
        return jsonify({"success": False})

    finally:
        conn.close()
        
@app.route('/delete-plan/<int:id>')
def delete_plan(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # delete plan
    cursor.execute("DELETE FROM plans WHERE id = ?", (id,))

    conn.commit()
    conn.close()

    flash("Plan deleted successfully", "success")
    return redirect(url_for('admin'))






init_excel()
# ---------------- Run App ---------------- #
if __name__ == '__main__':
    app.run(debug=True)
